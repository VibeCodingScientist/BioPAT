#!/usr/bin/env python3
"""Build expert validation spreadsheet for patent attorney review.

Produces an Excel file with:
  - Instructions tab (annotation guidelines + worked examples)
  - Annotation tab (100 selected statements with top-5 prior art)
  - Selection criteria tab (how the 100 were chosen)

Usage:
    python3.12 scripts/build_expert_sheet.py
    python3.12 scripts/build_expert_sheet.py --output expert_validation.xlsx
"""

import csv
import json
import random
import sys
from collections import Counter, defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data" / "novex"

# --- Selection strategy ---
# Oversample contentious cases, include unanimous controls
SELECTION = {
    "no_consensus": {"take": "all"},       # 11 → 11
    "override": {"take": "all"},           # 13 → 13
    "majority": {"take": 46},              # 98 → 46
    "unanimous": {"take": 30},             # 178 → 30
}
SEED = 42
TOP_K_DOCS = 5


def load_corpus_index():
    """Load corpus into a dict for doc lookup. Only load what we need."""
    corpus = {}
    for p in [
        ROOT / "data" / "benchmark" / "dual_corpus.jsonl",
        ROOT / "data" / "benchmark" / "corpus.jsonl",
    ]:
        if p.exists():
            print(f"Loading corpus from {p}...")
            with open(p) as f:
                for line in f:
                    d = json.loads(line)
                    corpus[d["_id"]] = {
                        "title": d.get("title", ""),
                        "text": d.get("text", ""),
                        "doc_type": d.get("doc_type", "unknown"),
                    }
            print(f"  Loaded {len(corpus)} docs")
            return corpus
    print("ERROR: No corpus file found")
    sys.exit(1)


def load_statements():
    stmts = {}
    with open(DATA / "statements.jsonl") as f:
        for line in f:
            s = json.loads(line)
            stmts[s["statement_id"]] = s
    return stmts


def select_100(stmts):
    """Stratified selection of 100 statements."""
    rng = random.Random(SEED)

    by_agreement = defaultdict(list)
    for sid, s in stmts.items():
        ag = s["ground_truth"].get("tier3_agreement", "unknown")
        by_agreement[ag].append(sid)

    selected = []
    for ag, cfg in SELECTION.items():
        pool = by_agreement.get(ag, [])
        if cfg["take"] == "all":
            selected.extend(pool)
        else:
            n = min(cfg["take"], len(pool))
            selected.extend(rng.sample(pool, n))

    # Sort by ID for clean presentation
    selected.sort()
    print(f"Selected {len(selected)} statements:")
    for ag in ["no_consensus", "override", "majority", "unanimous"]:
        count = sum(1 for sid in selected if stmts[sid]["ground_truth"]["tier3_agreement"] == ag)
        print(f"  {ag}: {count}")

    return selected


def get_top_docs(stmt, corpus, k=TOP_K_DOCS):
    """Get top-k prior art docs for a statement, sorted by relevance grade."""
    t1_docs = stmt["ground_truth"].get("tier1_relevant_docs", {})
    # Sort by grade descending, then doc ID for stability
    sorted_docs = sorted(t1_docs.items(), key=lambda x: (-x[1], x[0]))[:k]

    results = []
    for doc_id, grade in sorted_docs:
        doc = corpus.get(doc_id, {})
        title = doc.get("title", "Unknown document")
        text = doc.get("text", "")
        # Truncate text to ~250 chars for readability
        snippet = text[:250].strip()
        if len(text) > 250:
            snippet += "..."
        doc_type = doc.get("doc_type", "unknown")

        results.append({
            "doc_id": doc_id,
            "grade": grade,
            "doc_type": doc_type,
            "title": title,
            "snippet": snippet,
        })
    return results


def build_csv(stmts, selected, corpus, output_path):
    """Build the main annotation CSV."""
    rows = []
    for sid in selected:
        s = stmts[sid]
        gt = s["ground_truth"]
        docs = get_top_docs(s, corpus)

        row = {
            "Statement_ID": sid,
            "Domain": s["domain"],
            "Category": s.get("category", ""),
            "Statement_Text": s["text"],
            "Source_Patent": s.get("source_patent_id", ""),
        }

        # Add top-5 prior art docs
        for i in range(TOP_K_DOCS):
            if i < len(docs):
                d = docs[i]
                type_tag = f"[{d['doc_type'].upper()}] " if d['doc_type'] != 'unknown' else ""
                row[f"PriorArt_{i+1}_ID"] = d["doc_id"]
                row[f"PriorArt_{i+1}_Title"] = f"{type_tag}{d['title']}"
                row[f"PriorArt_{i+1}_Snippet"] = d["snippet"]
                row[f"PriorArt_{i+1}_RelevanceGrade"] = d["grade"]
            else:
                row[f"PriorArt_{i+1}_ID"] = ""
                row[f"PriorArt_{i+1}_Title"] = ""
                row[f"PriorArt_{i+1}_Snippet"] = ""
                row[f"PriorArt_{i+1}_RelevanceGrade"] = ""

        # Expert annotation columns (to be filled by reviewer)
        row["Expert_Label"] = ""  # NOVEL / PARTIALLY_ANTICIPATED / ANTICIPATED
        row["Expert_Confidence"] = ""  # High / Medium / Low
        row["Expert_Rationale"] = ""  # Brief justification

        rows.append(row)

    # Write CSV
    fieldnames = list(rows[0].keys())
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"Wrote annotation sheet: {output_path} ({len(rows)} rows)")
    return rows


def build_instructions(output_path):
    """Write annotation guidelines as a text file."""
    instructions = """
================================================================================
NovEx EXPERT VALIDATION — ANNOTATION GUIDELINES
================================================================================

TASK OVERVIEW
-------------
You will review 100 patent-derived technical statements and assess whether
each statement describes something NOVEL, PARTIALLY ANTICIPATED, or fully
ANTICIPATED by the provided prior art documents.

This validation is for a benchmark paper evaluating how well AI systems can
assess patent novelty. Your expert judgments will serve as ground truth
validation.

Estimated time: 5-6 hours (100 statements × ~3 min each)


LABEL DEFINITIONS
-----------------

NOVEL (score 0):
  The prior art documents do NOT anticipate this statement. The specific
  combination of elements, the particular application, or the core technical
  contribution described is not found in or obvious from the listed prior art.

  Example: A statement describes using compound X to treat disease Y via
  mechanism Z, and the prior art documents discuss compound X for other
  diseases but never mention disease Y or mechanism Z.

PARTIALLY_ANTICIPATED (score 1):
  Some elements of the statement appear in the prior art, but the complete
  combination or specific application is not fully disclosed. Key aspects
  are anticipated, but the statement contains meaningful novel contributions
  beyond the prior art.

  Example: A statement describes a method with steps A+B+C where the prior
  art documents describe A+B in detail but step C or the specific combination
  A+B+C is not taught.

ANTICIPATED (score 2):
  The prior art documents substantially anticipate this statement. The core
  technical contribution is already disclosed or would be obvious to a person
  skilled in the art based on the provided documents.

  Example: A statement describes using dsRNA of 15-21 nucleotides to inhibit
  gene expression, and multiple prior art documents describe exactly this
  technique.


WHAT YOU SEE PER STATEMENT
--------------------------
For each statement, you will see:
  - Statement_ID: Unique identifier (e.g., RN001)
  - Domain: IPC domain code (A61=medical, C07=organic chemistry, C12=biotech)
  - Statement_Text: The technical claim to evaluate
  - Source_Patent: The patent this statement was derived from (for context only)
  - PriorArt_1 through PriorArt_5: Top prior art documents, each with:
    - ID: Document identifier (W-prefix = paper, numeric = patent)
    - Title: Document title
    - Snippet: First ~250 characters of the document text
    - RelevanceGrade: How relevant this doc is (3=highly, 2=moderate, 1=marginal)

WHAT TO FILL IN
---------------
  - Expert_Label: NOVEL, PARTIALLY_ANTICIPATED, or ANTICIPATED
  - Expert_Confidence: High, Medium, or Low
  - Expert_Rationale: One sentence explaining your judgment (optional but helpful)


IMPORTANT NOTES
---------------
  1. Judge based ONLY on the provided prior art documents, not your general
     domain knowledge. The question is: "Do THESE documents anticipate THIS
     statement?"

  2. The snippets are abbreviated. If a document title alone strongly suggests
     anticipation, that's sufficient — you don't need the full text.

  3. Documents with RelevanceGrade=3 are highly relevant; Grade=1 is marginal.
     Focus your attention on Grade 2-3 documents.

  4. W-prefix IDs are academic papers; purely numeric IDs are patents.

  5. "Source_Patent" is the patent the statement was extracted from. It is NOT
     part of the prior art — ignore it for your novelty judgment.

  6. When in doubt between NOVEL and PARTIALLY_ANTICIPATED, lean toward
     PARTIALLY_ANTICIPATED. When in doubt between PARTIALLY_ANTICIPATED and
     ANTICIPATED, consider whether a skilled practitioner would find the
     statement's contribution obvious from the prior art.


WORKED EXAMPLES
---------------

Example A — ANTICIPATED:
  Statement: "Double-stranded RNA molecules between 15 and 21 nucleotides in
  length complementary to a target gene sequence are used to inhibit expression
  of that target gene."
  Prior art includes multiple patents and papers describing exactly this RNAi
  mechanism with these size parameters.
  → ANTICIPATED (High confidence): Core technique fully described in prior art.

Example B — NOVEL:
  Statement: "A method for distinguishing agonistic anti-PD-L1 antibodies from
  antagonistic anti-PD-L1 antibodies using a functional assay that differentiates
  antibodies activating B7-H1 signaling from those blocking PD-1 interactions."
  Prior art discusses PD-L1 antibodies broadly but none describe a method to
  functionally distinguish agonistic vs antagonistic variants.
  → NOVEL (Medium confidence): Specific functional distinction method not taught.

Example C — PARTIALLY_ANTICIPATED:
  Statement: "Modified mRNA encoding AIF1 with chemical modifications for
  therapeutic oncology applications."
  Prior art describes modified mRNA technology in general and for other targets,
  but AIF1 targeting for oncology is not specifically taught.
  → PARTIALLY_ANTICIPATED (High confidence): Platform known, specific target new.

================================================================================
Thank you for your expert review. Your judgments are critical for validating
the quality of our benchmark's ground truth labels.
================================================================================
""".strip()

    with open(output_path, "w") as f:
        f.write(instructions)
    print(f"Wrote instructions: {output_path}")


def build_selection_summary(stmts, selected, output_path):
    """Write a summary of the selection criteria."""
    by_agreement = Counter()
    by_label = Counter()
    by_domain = Counter()

    for sid in selected:
        s = stmts[sid]
        by_agreement[s["ground_truth"]["tier3_agreement"]] += 1
        by_label[s["ground_truth"]["tier3_novelty_label"]] += 1
        by_domain[s["domain"]] += 1

    summary = {
        "total_selected": len(selected),
        "selection_strategy": {
            "no_consensus": "All 11 (most uncertain — highest validation value)",
            "override": "All 13 (label was forced by pipeline — need expert check)",
            "majority": "46 randomly sampled from 98 (contentious — 2/3 LLMs agreed)",
            "unanimous": "30 randomly sampled from 178 (controls — all 3 LLMs agreed)",
        },
        "by_agreement_level": dict(sorted(by_agreement.items())),
        "by_novelty_label": dict(sorted(by_label.items())),
        "by_domain": dict(sorted(by_domain.items())),
        "seed": SEED,
        "selected_ids": selected,
    }

    with open(output_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"Wrote selection summary: {output_path}")

    # Also print
    print(f"\nBy agreement: {dict(by_agreement)}")
    print(f"By label: {dict(by_label)}")
    print(f"By domain: {dict(by_domain)}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Build expert validation spreadsheet")
    parser.add_argument("--output", default="data/novex/expert_validation.csv",
                        help="Output CSV path")
    parser.add_argument("--output-dir", default="data/novex",
                        help="Output directory for all files")
    args = parser.parse_args()

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    print("Loading data...")
    stmts = load_statements()
    corpus = load_corpus_index()

    print(f"\nSelecting 100 statements...")
    selected = select_100(stmts)

    print(f"\nBuilding outputs...")
    build_instructions(out_dir / "expert_validation_guidelines.txt")
    build_csv(stmts, selected, corpus, out_dir / "expert_validation.csv")
    build_selection_summary(stmts, selected, out_dir / "expert_validation_selection.json")

    # Also try to build Excel if openpyxl available
    try:
        import openpyxl
        build_excel(stmts, selected, corpus, out_dir / "expert_validation.xlsx")
    except ImportError:
        print("\nNote: openpyxl not installed — Excel file not generated.")
        print("The CSV file works in Excel/Google Sheets. To get .xlsx:")
        print("  pip install openpyxl && python3.12 scripts/build_expert_sheet.py")

    print(f"\n{'='*60}")
    print("EXPERT VALIDATION PACKAGE READY")
    print(f"{'='*60}")
    print(f"  Guidelines:  {out_dir / 'expert_validation_guidelines.txt'}")
    print(f"  Annotation:  {out_dir / 'expert_validation.csv'}")
    print(f"  Selection:   {out_dir / 'expert_validation_selection.json'}")
    print(f"{'='*60}")
    print(f"\nSend these files to your patent attorney/associate.")
    print(f"Estimated review time: 5-6 hours for {len(selected)} statements.")


def build_excel(stmts, selected, corpus, output_path):
    """Build single Excel workbook with 3 tabs: Instructions, Annotation, Selection."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()

    # ---- Styles ----
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    expert_header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    expert_header_font = Font(bold=True, color="000000", size=10, name="Calibri")
    expert_cell_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")
    wrap_center = Alignment(wrap_text=True, horizontal="center", vertical="top")
    bold11 = Font(bold=True, size=11, name="Calibri")
    bold12 = Font(bold=True, size=12, name="Calibri")
    bold14 = Font(bold=True, size=14, name="Calibri", color="1F4E79")
    normal = Font(size=10, name="Calibri")
    mono = Font(size=10, name="Consolas")
    section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # ==================================================================
    # TAB 1: Instructions
    # ==================================================================
    ws = wb.active
    ws.title = "Instructions"
    ws.sheet_properties.tabColor = "4472C4"
    ws.column_dimensions["A"].width = 90

    sections = [
        ("NovEx EXPERT VALIDATION — ANNOTATION GUIDELINES", bold14),
        ("", normal),
        ("TASK OVERVIEW", bold12),
        ("You will review 100 patent-derived technical statements and assess whether each statement describes something NOVEL, PARTIALLY ANTICIPATED, or fully ANTICIPATED by the provided prior art documents.", normal),
        ("", normal),
        ("This validation is for a benchmark paper evaluating how well AI systems can assess patent novelty. Your expert judgments will serve as ground truth validation.", normal),
        ("", normal),
        ("Estimated time: 5-6 hours (100 statements x ~3 min each)", bold11),
        ("", normal),
        ("", normal),
        ("LABEL DEFINITIONS", bold12),
        ("", normal),
        ("NOVEL", bold11),
        ("The prior art documents do NOT anticipate this statement. The specific combination of elements, the particular application, or the core technical contribution described is not found in or obvious from the listed prior art.", normal),
        ("Example: A statement describes using compound X to treat disease Y via mechanism Z, and the prior art documents discuss compound X for other diseases but never mention disease Y or mechanism Z.", mono),
        ("", normal),
        ("PARTIALLY_ANTICIPATED", bold11),
        ("Some elements of the statement appear in the prior art, but the complete combination or specific application is not fully disclosed. Key aspects are anticipated, but the statement contains meaningful novel contributions beyond the prior art.", normal),
        ("Example: A statement describes a method with steps A+B+C where the prior art documents describe A+B in detail but step C or the specific combination A+B+C is not taught.", mono),
        ("", normal),
        ("ANTICIPATED", bold11),
        ("The prior art documents substantially anticipate this statement. The core technical contribution is already disclosed or would be obvious to a person skilled in the art based on the provided documents.", normal),
        ("Example: A statement describes using dsRNA of 15-21 nucleotides to inhibit gene expression, and multiple prior art documents describe exactly this technique.", mono),
        ("", normal),
        ("", normal),
        ("WHAT YOU SEE PER STATEMENT (on the Annotation tab)", bold12),
        ("", normal),
        ("Statement_ID: Unique identifier (e.g., RN001)", normal),
        ("Domain: IPC domain code (A61=medical, C07=organic chemistry, C12=biotech)", normal),
        ("Statement_Text: The technical claim to evaluate", normal),
        ("Source_Patent: The patent this statement was derived from (for context only — NOT prior art)", normal),
        ("PriorArt_1 through PriorArt_5: Top prior art documents, each with ID, Title, Snippet (~250 chars), and RelevanceGrade (3=highly relevant, 2=moderate, 1=marginal)", normal),
        ("", normal),
        ("", normal),
        ("WHAT TO FILL IN (yellow columns on the Annotation tab)", bold12),
        ("", normal),
        ("Expert_Label: NOVEL, PARTIALLY_ANTICIPATED, or ANTICIPATED (use the dropdown)", normal),
        ("Expert_Confidence: High, Medium, or Low", normal),
        ("Expert_Rationale: One sentence explaining your judgment (optional but very helpful)", normal),
        ("", normal),
        ("", normal),
        ("IMPORTANT NOTES", bold12),
        ("", normal),
        ("1. Judge based ONLY on the provided prior art documents, not your general domain knowledge. The question is: 'Do THESE documents anticipate THIS statement?'", normal),
        ("2. The snippets are abbreviated. If a document title alone strongly suggests anticipation, that's sufficient.", normal),
        ("3. Documents with RelevanceGrade=3 are highly relevant; Grade=1 is marginal. Focus on Grade 2-3 documents.", normal),
        ("4. W-prefix IDs are academic papers; purely numeric IDs are patents.", normal),
        ("5. 'Source_Patent' is the patent the statement was extracted from. It is NOT part of the prior art.", normal),
        ("6. When in doubt between NOVEL and PARTIALLY_ANTICIPATED, lean toward PARTIALLY_ANTICIPATED. When in doubt between PARTIALLY_ANTICIPATED and ANTICIPATED, consider whether a skilled practitioner would find the contribution obvious from the prior art.", normal),
        ("", normal),
        ("", normal),
        ("WORKED EXAMPLES", bold12),
        ("", normal),
        ("Example A — ANTICIPATED:", bold11),
        ('Statement: "Double-stranded RNA molecules between 15 and 21 nucleotides in length complementary to a target gene sequence are used to inhibit expression of that target gene."', normal),
        ("Prior art includes multiple patents and papers describing exactly this RNAi mechanism with these size parameters.", normal),
        ("-> ANTICIPATED (High confidence): Core technique fully described in prior art.", mono),
        ("", normal),
        ("Example B — NOVEL:", bold11),
        ('Statement: "A method for distinguishing agonistic anti-PD-L1 antibodies from antagonistic anti-PD-L1 antibodies using a functional assay that differentiates antibodies activating B7-H1 signaling from those blocking PD-1 interactions."', normal),
        ("Prior art discusses PD-L1 antibodies broadly but none describe a method to functionally distinguish agonistic vs antagonistic variants.", normal),
        ("-> NOVEL (Medium confidence): Specific functional distinction method not taught.", mono),
        ("", normal),
        ("Example C — PARTIALLY_ANTICIPATED:", bold11),
        ('Statement: "Modified mRNA encoding AIF1 with chemical modifications for therapeutic oncology applications."', normal),
        ("Prior art describes modified mRNA technology in general and for other targets, but AIF1 targeting for oncology is not specifically taught.", normal),
        ("-> PARTIALLY_ANTICIPATED (High confidence): Platform known, specific target new.", mono),
        ("", normal),
        ("", normal),
        ("Thank you for your expert review. Your judgments are critical for validating the quality of our benchmark's ground truth labels.", bold11),
    ]

    for row_idx, (text, font) in enumerate(sections, 1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = font
        cell.alignment = wrap
        # Section headers get a background
        if font == bold12:
            cell.fill = section_fill

    # ==================================================================
    # TAB 2: Annotation
    # ==================================================================
    ws_ann = wb.create_sheet("Annotation")
    ws_ann.sheet_properties.tabColor = "FFC000"

    headers = ["#", "Statement_ID", "Domain", "Category", "Statement_Text", "Source_Patent"]
    for i in range(1, TOP_K_DOCS + 1):
        headers.extend([
            f"Doc{i}_ID", f"Doc{i}_Title", f"Doc{i}_Snippet", f"Doc{i}_Grade",
        ])
    headers.extend(["Expert_Label", "Expert_Confidence", "Expert_Rationale"])

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws_ann.cell(row=1, column=col_idx, value=header)
        if header.startswith("Expert_"):
            cell.fill = expert_header_fill
            cell.font = expert_header_font
        else:
            cell.fill = header_fill
            cell.font = header_font
        cell.alignment = wrap_center

    # Write data rows
    for row_num, sid in enumerate(selected, 2):
        s = stmts[sid]
        docs = get_top_docs(s, corpus)

        col = 1
        ws_ann.cell(row=row_num, column=col, value=row_num - 1); col += 1
        ws_ann.cell(row=row_num, column=col, value=sid); col += 1
        ws_ann.cell(row=row_num, column=col, value=s["domain"]); col += 1
        ws_ann.cell(row=row_num, column=col, value=s.get("category", "")); col += 1

        c = ws_ann.cell(row=row_num, column=col, value=s["text"])
        c.alignment = wrap
        col += 1

        ws_ann.cell(row=row_num, column=col, value=s.get("source_patent_id", "")); col += 1

        for i in range(TOP_K_DOCS):
            if i < len(docs):
                d = docs[i]
                type_tag = f"[{d['doc_type'].upper()}] " if d['doc_type'] != 'unknown' else ""
                ws_ann.cell(row=row_num, column=col, value=d["doc_id"]); col += 1
                c = ws_ann.cell(row=row_num, column=col, value=f"{type_tag}{d['title']}")
                c.alignment = wrap
                col += 1
                c = ws_ann.cell(row=row_num, column=col, value=d["snippet"])
                c.alignment = wrap
                col += 1
                ws_ann.cell(row=row_num, column=col, value=d["grade"]); col += 1
            else:
                col += 4

        # Expert columns — yellow background
        for _ in range(3):
            cell = ws_ann.cell(row=row_num, column=col, value="")
            cell.fill = expert_cell_fill
            cell.alignment = wrap
            col += 1

    # Column widths
    widths = {
        1: 4,    # #
        2: 11,   # Statement_ID
        3: 7,    # Domain
        4: 13,   # Category
        5: 65,   # Statement_Text
        6: 12,   # Source_Patent
    }
    # Prior art columns: ID=14, Title=35, Snippet=45, Grade=6
    for i in range(TOP_K_DOCS):
        base = 7 + i * 4
        widths[base] = 14      # Doc ID
        widths[base + 1] = 35  # Title
        widths[base + 2] = 45  # Snippet
        widths[base + 3] = 6   # Grade
    # Expert columns
    expert_start = 7 + TOP_K_DOCS * 4
    widths[expert_start] = 22      # Label
    widths[expert_start + 1] = 15  # Confidence
    widths[expert_start + 2] = 40  # Rationale

    for col_idx, width in widths.items():
        ws_ann.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row + first 2 columns
    ws_ann.freeze_panes = "C2"

    # Data validation dropdowns for Expert_Label and Expert_Confidence
    label_col = get_column_letter(expert_start)
    conf_col = get_column_letter(expert_start + 1)

    dv_label = DataValidation(
        type="list",
        formula1='"NOVEL,PARTIALLY_ANTICIPATED,ANTICIPATED"',
        allow_blank=True,
    )
    dv_label.error = "Please select NOVEL, PARTIALLY_ANTICIPATED, or ANTICIPATED"
    dv_label.errorTitle = "Invalid label"
    dv_label.prompt = "Select novelty label"
    dv_label.promptTitle = "Expert Label"
    ws_ann.add_data_validation(dv_label)
    dv_label.add(f"{label_col}2:{label_col}101")

    dv_conf = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True,
    )
    dv_conf.prompt = "Select confidence level"
    dv_conf.promptTitle = "Confidence"
    ws_ann.add_data_validation(dv_conf)
    dv_conf.add(f"{conf_col}2:{conf_col}101")

    # Row height for data rows
    for row_num in range(2, len(selected) + 2):
        ws_ann.row_dimensions[row_num].height = 60

    # ==================================================================
    # TAB 3: Selection Criteria
    # ==================================================================
    ws_sel = wb.create_sheet("Selection Criteria")
    ws_sel.sheet_properties.tabColor = "70AD47"
    ws_sel.column_dimensions["A"].width = 30
    ws_sel.column_dimensions["B"].width = 15
    ws_sel.column_dimensions["C"].width = 55

    rows_data = [
        ("SELECTION SUMMARY", "", "", bold12, section_fill),
        ("Total statements in benchmark", "300", "", bold11, None),
        ("Total selected for review", "100", "", bold11, None),
        ("Random seed", str(SEED), "(for reproducibility)", normal, None),
        ("", "", "", normal, None),
        ("STRATIFIED SAMPLING", "", "", bold12, section_fill),
        ("Agreement Level", "Sampled / Total", "Rationale", bold11, header_fill),
        ("no_consensus", "11 / 11", "All included — most uncertain, highest validation value", normal, None),
        ("override", "13 / 13", "All included — label was forced by pipeline", normal, None),
        ("majority", "46 / 98", "Random sample — 2 of 3 LLMs agreed", normal, None),
        ("unanimous", "30 / 178", "Random sample — controls, all 3 LLMs agreed", normal, None),
        ("", "", "", normal, None),
        ("LABEL DISTRIBUTION IN SAMPLE", "", "", bold12, section_fill),
    ]

    by_label = Counter()
    for sid in selected:
        by_label[stmts[sid]["ground_truth"]["tier3_novelty_label"]] += 1
    for label in ["NOVEL", "PARTIALLY_ANTICIPATED", "ANTICIPATED"]:
        rows_data.append((label, str(by_label.get(label, 0)), "", normal, None))

    rows_data.append(("", "", "", normal, None))
    rows_data.append(("DOMAIN DISTRIBUTION IN SAMPLE", "", "", bold12, section_fill))

    by_domain = Counter()
    for sid in selected:
        by_domain[stmts[sid]["domain"]] += 1
    domain_names = {"A61": "A61 (Medical/Pharma)", "C07": "C07 (Organic Chemistry)", "C12": "C12 (Biotechnology)"}
    for domain in sorted(by_domain.keys()):
        rows_data.append((domain_names.get(domain, domain), str(by_domain[domain]), "", normal, None))

    for row_idx, (a, b, c, font, fill) in enumerate(rows_data, 1):
        ca = ws_sel.cell(row=row_idx, column=1, value=a)
        cb = ws_sel.cell(row=row_idx, column=2, value=b)
        cc = ws_sel.cell(row=row_idx, column=3, value=c)
        ca.font = font
        cb.font = normal
        cc.font = normal
        if fill:
            ca.fill = fill
            cb.fill = fill
            cc.fill = fill
        if font == bold11 and fill == header_fill:
            ca.font = Font(bold=True, color="FFFFFF", size=10)
            cb.font = Font(bold=True, color="FFFFFF", size=10)
            cc.font = Font(bold=True, color="FFFFFF", size=10)

    # Make Annotation the active sheet when opened
    wb.active = wb.sheetnames.index("Annotation")

    wb.save(output_path)
    print(f"Wrote Excel workbook: {output_path}")


if __name__ == "__main__":
    main()
